# This file is part of STAAS-Reporting-Fusion.
#
# STAAS-Reporting-Fusion is licensed under the BSD 2-Clause License.
# You may obtain a copy of the License at
#
#     https://opensource.org/licenses/BSD-2-Clause
#
# Redistribution and use in source and binary forms, with or without
# modification, are permitted provided that the following conditions are met:
#
# 1. Redistributions of source code must retain the above copyright notice,
#    this list of conditions and the following disclaimer.
# 2. Redistributions in binary form must reproduce the above copyright notice,
#    this list of conditions and the following disclaimer in the documentation
#    and/or other materials provided with the distribution.
#
# THIS SOFTWARE IS PROVIDED BY THE COPYRIGHT HOLDERS AND CONTRIBUTORS "AS IS"
# AND ANY EXPRESS OR IMPLIED WARRANTIES, INCLUDING, BUT NOT LIMITED TO, THE
# IMPLIED WARRANTIES OF MERCHANTABILITY AND FITNESS FOR A PARTICULAR PURPOSE
# ARE DISCLAIMED. IN NO EVENT SHALL THE COPYRIGHT HOLDER OR CONTRIBUTORS BE
# LIABLE FOR ANY DIRECT, INDIRECT, INCIDENTAL, SPECIAL, EXEMPLARY, OR
# CONSEQUENTIAL DAMAGES (INCLUDING, BUT NOT LIMITED TO, PROCUREMENT OF
# SUBSTITUTE GOODS OR SERVICES; LOSS OF USE, DATA, OR PROFITS; OR BUSINESS
# INTERRUPTION) HOWEVER CAUSED AND ON ANY THEORY OF LIABILITY, WHETHER IN
# CONTRACT, STRICT LIABILITY, OR TORT (INCLUDING NEGLIGENCE OR OTHERWISE)
# ARISING IN ANY WAY OUT OF THE USE OF THIS SOFTWARE, EVEN IF ADVISED OF THE
# POSSIBILITY OF SUCH DAMAGE.

import os
import pandas as pd
import urllib3
import pprint as pp
from datetime import datetime
from openpyxl import load_workbook
from packaging.version import Version
from staas_common import (
    parse_arguments,
    initialise_client,
    check_purity_role,
    check_api_version,
    list_fleets,
    list_members,
    pypureclient
)


debug = 3
REALMS_VERSION = "1.66"

# Header rows for the reporting spreadsheet are defined here
VOLUME_HEADER_ROWS = [
    ['Date/Time', 'Array', 'Volume']
]

ARRAY_HEADER_ROWS = [
    ['Date/Time', 'Array']  # Initial headers, will be updated dynamically
]

REALM_HEADER_ROWS = [
    ['Date/Time', 'Array', 'Realm']
]

DIRECTORY_HEADER_ROWS = [
    ['Date/Time', 'Array', 'Directory']
]

# Disable certificate warnings
urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)

def get_volume_space(client, fleet_member_name, volumes):
    space_values = {}
    chunk_size = 500

    for i in range(0, len(volumes), chunk_size):
        volume_chunk = volumes[i:i + chunk_size]
        response = client.get_volumes_space(context_names=fleet_member_name, names=volume_chunk)
        if response.status_code == 200:
            if debug >= 4:
                print(f"Space values for volumes in array {fleet_member_name}")
            for volume in response.items:
                space_values[volume.name] = volume.space.__dict__
        else:
            print(f"Failed to retrieve volume space. Status code: {response.status_code}, Error: {response.errors}")
            break

    return space_values

def read_volume_tags(client, fleet_member, volumes):
    tags = {}
    chunk_size = 500

    for i in range(0, len(volumes), chunk_size):
        volume_chunk = volumes[i:i + chunk_size]
        response = client.get_volumes_tags(context_names=[fleet_member], resource_names=volume_chunk, namespaces=NAMESPACE)
        # Check the response
        if response.status_code == 200:
            if debug >= 4:
                print(f"Tags from {fleet_member} in namespace {NAMESPACE}:")
                pp.pprint(response.items)
            # Process each tag, which has a context, key, value, namespace, and resource (of a volume)
            for tag in response.items:
                if tag.namespace == NAMESPACE and tag.key == TAG_KEY:
                    volume = tag.resource.name
                    tags[volume] = tag.value
        else:
            print(f"Failed to get tags from {fleet_member}. Status code: {response.status_code}, Error: {response.errors}")
            break

    return tags

def report_volumes(client, fleet_member, namespace, tag_key):
    volume_set = []
    response = client.get_volumes(context_names=[fleet_member])
    if response.status_code == 200:
        if debug >= 2:
            print(f"Finding volumes for array {fleet_member}")
        for volume in response.items:
            if volume.subtype == 'regular':
                volume_set.append(volume.name)
            else:
                if debug >= 4:
                    print(f'Non-regular volume {volume.name} found - not reporting it.')

        # Retrieve tags and space values for the volumes
        tags = read_volume_tags(client, fleet_member, volume_set)
        space_values = get_volume_space(client, fleet_member, volume_set)

        # Initialize the volumes_by_tag dictionary
        volumes_by_tag = {}

        # Dynamically update headers if space values exist
        if space_values:
            first_space = next(iter(space_values.values()))
            headers = ['Date/Time', 'Array', 'Volume'] + list(first_space.keys())
            VOLUME_HEADER_ROWS[0] = headers

        # Process each volume
        for volume in volume_set:
            tag = tags.get(volume, 'NoChargebackTag')  # Default to 'NoChargebackTag' if no tag is found
            space = space_values.get(volume, {})
            volume_info = {
                'Date/Time': NOW,
                'Array': fleet_member,
                'Volume': volume
            }

            # Add space attributes dynamically
            for key in VOLUME_HEADER_ROWS[0][3:]:
                volume_info[key] = space.get(key, '')

            # Add the volume to the appropriate tag group
            if tag not in volumes_by_tag:
                volumes_by_tag[tag] = []
            volumes_by_tag[tag].append(volume_info)

        return volumes_by_tag
    else:
        print(f"Failed to retrieve volumes. Status code: {response.status_code}, Error: {response.errors}")
        return {}

def report_arrays(client, fleet, fleet_members):
    fleet_space_report = {}
    realm_space_report = {}

    for fleet_member in fleet_members:
        # Report array space usage
        response = client.get_arrays_space(context_names=[fleet_member])
        if response.status_code == 200:
            if debug >= 2:
                print(f"Space usage for array {fleet_member}")
            for item in response.items:
                if hasattr(item, 'space'):
                    space = item.space.__dict__
                    space_report = {'Date/Time': NOW, 'Array': fleet_member}
                    space_report.update(space)
                    if fleet_member not in fleet_space_report:
                        fleet_space_report[fleet_member] = []
                    fleet_space_report[fleet_member].append(space_report)
                else:
                    print(f"No space information available for array {fleet_member}")
        else:
            print(f"Failed to retrieve space usage. Status code: {response.status_code}, Error: {response.errors}")

        # Check API version before reporting realm space usage
        if Version(pypureclient.__version__) >= Version(REALMS_VERSION):
            connection_type_response = client.get_fleets_members()
            if connection_type_response.status_code == 200:
                for member in connection_type_response.items:
                    if member.member.name == fleet_member:  # and member.member.is_local == 'true':
                        # Report realm space usage
                        response = client.get_realms_space(context_names=[fleet_member])
                        if response.status_code == 200:
                            if debug >= 2:
                                print(f"Space usage for realms in array {fleet_member}")
                            for item in response.items:
                                if hasattr(item, 'space'):
                                    space = item.space.__dict__
                                    space_report = {'Date/Time': NOW, 'Array': fleet_member, 'Realm': item.name}
                                    space_report.update(space)
                                    if item.name not in realm_space_report:
                                        realm_space_report[item.name] = []
                                    realm_space_report[item.name].append(space_report)
                                else:
                                    print(f"No space information available for realms in array {fleet_member}")
                        else:
                            print(f"Failed to retrieve realm space usage. Status code: {response.status_code}, Error: {response.errors}")
            else:
                print(f"Failed to retrieve connection type. Status code: {connection_type_response.status_code}, Error: {connection_type_response.errors}")
        else:
            if debug >=2:
                print("Skipping realm space report as pypureclient version is too low")

    return fleet_space_report, realm_space_report

def report_directories(client, fleet_member):
    directory_set = []
    limit = 200  # Number of directories to process per request
    continuation_token = None  # Token for paginated requests

    while True:
        # Retrieve directories in chunks of 200
        if response.status_code == 200:
            if debug >= 2:
                print(f"Finding directories for array {fleet_member} (Batch with continuation token: {continuation_token})")
            for directory in response.items:
                directory_info = {
                    'Date/Time': NOW,
                    'Array': fleet_member,
                    'Directory': directory.name
                }
                # Dynamically add all attributes from the space object
                if hasattr(directory, 'space') and directory.space:
                    directory_info.update(directory.space.__dict__)
                else:
                    print(f"No space information available for directory {directory.name}")
                directory_set.append(directory_info)

                # Dynamically update DIRECTORY_HEADER_ROWS based on the first directory
                if len(DIRECTORY_HEADER_ROWS[0]) == 3:  # Only update if headers are not already updated
                    additional_headers = [key for key in directory_info.keys() if key not in DIRECTORY_HEADER_ROWS[0]]
                    DIRECTORY_HEADER_ROWS[0].extend(additional_headers)

            # Check if there are more directories to process
            continuation_token = response.continuation_token
            if not continuation_token:
                break  # Exit the loop if there are no more directories
        else:
            print(f"Failed to retrieve directories. Status code: {response.status_code}, Error: {response.errors}")
            break

    return directory_set

def save_report_to_excel(report_data, headers, report_path, sheet_prefix):
    try:
        # Check if the file exists
        if os.path.exists(report_path):
            try:
                book = load_workbook(report_path)
                with pd.ExcelWriter(report_path, engine='openpyxl', mode='a', if_sheet_exists='overlay') as writer:
                    for group, data in report_data.items():
                        if not data:
                            print(f"Warning: No data for group '{group}'. Skipping.")
                            continue
                        df = pd.DataFrame(data)
                        sheet_name = f"{sheet_prefix} {group}"
                        if sheet_name in book.sheetnames:
                            startrow = book[sheet_name].max_row
                            df.to_excel(writer, sheet_name=sheet_name, index=False, header=False, startrow=startrow)
                        else:
                            # Write headers when creating a new sheet
                            df.to_excel(writer, sheet_name=sheet_name, index=False, header=headers)
            except (KeyError, ValueError) as e:
                print(f"Error processing workbook: {e}. Creating a new file.")
                with pd.ExcelWriter(report_path, engine='openpyxl') as writer:
                    for group, data in report_data.items():
                        if not data:
                            continue
                        df = pd.DataFrame(data)
                        # Write headers when creating a new file
                        df.to_excel(writer, sheet_name=f"{sheet_prefix} {group}", index=False, header=headers)
        else:
            with pd.ExcelWriter(report_path, engine='openpyxl') as writer:
                for group, data in report_data.items():
                    if not data:
                        continue
                    df = pd.DataFrame(data)
                    # Write headers when creating a new file
                    df.to_excel(writer, sheet_name=f"{sheet_prefix} {group}", index=False, header=headers)
    except PermissionError as e:
        print(f"PermissionError: {e}. Please ensure the file is not open in another application.")
    except ValueError as e:
        print(f"ValueError: {e}. Please check the structure of report_data.")

# Main script
if __name__ == "__main__":
    # Parse command-line arguments
    args = parse_arguments("report")

    # Read the configuration file
    config_path = os.path.join(args.config)
    if not os.path.exists(config_path):
        print(f"Configuration file not found: {config_path}")
        exit(1)

    config_spreadsheet = pd.ExcelFile(config_path)

    # Read the Excel file
    fleet_df = config_spreadsheet.parse('Fleet')

    # Extract global variables from the Fleet sheet
    global_variables = fleet_df.iloc[0].to_dict()

    # Assign global variables
    USER_NAME = os.getenv('PURE_USER_NAME')
    API_TOKEN = os.getenv('PURE_API_TOKEN')
    FUSION_SERVER = global_variables.get('FUSION_SERVER', '')
    NAMESPACE = global_variables.get('NAMESPACE', '')
    TAG_KEY = "chargeback"

    MNTH = datetime.now().strftime("%Y-%m")
    NOW = datetime.now().strftime("%Y-%m-%d %H:%M")

    print(f"Connecting to Fusion server: {FUSION_SERVER} with user: {USER_NAME}")

    client = initialise_client(FUSION_SERVER, USER_NAME, API_TOKEN)
    if not client:
        exit(1)
    # Check to see minimum version of 2.41 & array admin privileges for this user
    role = check_purity_role(client, USER_NAME) 
    if not (role == "array_admin" or role == "read_only)"):
       exit(1)
    if not check_api_version(client, 2.41):
        exit(2)

    # Get the arrays for reporting contexts for the nominated fleet
    fleets = list_fleets(client)

    # At some point, there may be multiple fleets visible from a single fusion end-point
    for fleet in fleets:
        volume_space_report = {}
        directory_space_report = {}
        fleet_members = list_members(client, [fleet])

        for fleet_member in fleet_members:

            # Generate volume space report
            volumes_by_tag = report_volumes(client, fleet_member, NAMESPACE, TAG_KEY)
            for tag, volumes in volumes_by_tag.items():
                if tag not in volume_space_report:
                    volume_space_report[tag] = []
                volume_space_report[tag].extend(volumes)

            """
            # Generate directory space report, when supported in fusion
            if not fleet[fleet_member].is_local:
                if debug >= 2:
                    print(f"Skipping non-local fleet member: {fleet_member.name}")
                continue
            directories = report_directories(client, fleet_member)
            if fleet_member not in directory_space_report:
                directory_space_report[fleet_member] = []
            directory_space_report[fleet_member].extend(directories)
            """

        # Save the reports
        volumes_report_path = os.path.join(args.reportdir, f"Space-Report-Volumes-{MNTH}.xlsx")
        save_report_to_excel(volume_space_report, VOLUME_HEADER_ROWS[0], volumes_report_path, 'Tag')

        """
        # Generate directory space report, when supported in fusion
        directories_report_path = os.path.join(args.reportdir, f"Space-Report-Directories-{MNTH}.xlsx")
        save_report_to_excel(directory_space_report, DIRECTORY_HEADER_ROWS[0], directories_report_path, 'Directory')
        """