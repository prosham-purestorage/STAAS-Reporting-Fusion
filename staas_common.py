# filepath: staas-common.py
# This file is part of STAAS-Reporting-Fusion.
#
# STAAS-Reporting-Fusion is licensed under the BSD 2-Clause License.
# You may obtain a copy of the License at
#
#     https://opensource.org/licenses/BSD-2-Clause
#
# Redistribution and use in source and binary forms, with or without
# modification, are permitted provided that the following conditions are met:
#
# 1. Redistributions of source code must retain the above copyright notice,
#    this list of conditions and the following disclaimer.
# 2. Redistributions in binary form must reproduce the above copyright notice,
#    this list of conditions and the following disclaimer in the documentation
#    and/or other materials provided with the distribution.
#
# THIS SOFTWARE IS PROVIDED BY THE COPYRIGHT HOLDERS AND CONTRIBUTORS "AS IS"
# AND ANY EXPRESS OR IMPLIED WARRANTIES; WITHOUT LIMITATION, THE IMPLIED
# WARRANTIES OF MERCHANTABILITY AND FITNESS FOR A PARTICULAR PURPOSE ARE
# DISCLAIMED. IN NO EVENT SHALL THE COPYRIGHT HOLDER OR CONTRIBUTORS BE
# LIABLE FOR ANY DIRECT, INDIRECT, INCIDENTAL, SPECIAL, EXEMPLARY, OR
# CONSEQUENTIAL DAMAGES (INCLUDING, BUT NOT LIMITED TO, PROCUREMENT OF
# SUBSTITUTE GOODS OR SERVICES; LOSS OF USE, DATA, OR PROFITS; OR BUSINESS
# INTERRUPTION) HOWEVER CAUSED AND ON ANY THEORY OF LIABILITY, WHETHER IN
# CONTRACT, STRICT LIABILITY, OR TORT (INCLUDING NEGLIGENCE OR OTHERWISE)
# ARISING IN ANY WAY OUT OF THE USE OF THIS SOFTWARE, EVEN IF ADVISED OF THE
# POSSIBILITY OF SUCH DAMAGE.

import os
import pandas as pd
import pypureclient
import urllib3
import re
import pprint as pp

from openpyxl import load_workbook
from datetime import datetime
from pypureclient import flasharray
from pypureclient.flasharray import Client, PureError
from staas_common import check_purity_role, check_api_version, initialise_client, list_fleets

debug = 4

def read_volume_tags(fleet_member_name, volumes):
    tags = {}
    # Get any tags in the correct namespace for this array, and create a map of (array,volumes) to tags
    response = client.get_volumes_tags(context_names=fleet_member_name, resource_names=volumes, namespaces=NAMESPACE)
    # Check the response
    if response.status_code == 200:
        if debug >= 4:
            print(f"Tags from {fleet_member_name} in namespace {NAMESPACE}:")
            pp.pprint(response.items)
        # Process each tag, which has a context, key, value, namespace, and resource (of a volume)
        for tag in response.items:
            if tag.namespace == NAMESPACE and tag.key == TAG_KEY:
                volume = tag.resource.name
                tags[volume] = tag.value
    else:
        print(f"Failed to get tags from {fleet_member_name}. Status code: {response.status_code}, Error: {response.errors}")
    return tags

def get_volume_space(fleet_member_name, volumes):
    space_values = {}
    response = client.get_volumes_space(context_names=fleet_member_name, names=volumes)
    if response.status_code == 200:
        if debug >= 2:
            print(f"Space values for volumes in array {fleet_member_name}")
        for volume in response.items:
            space_values[volume.name] = volume.space.__dict__
    else:
        print(f"Failed to retrieve volume space. Status code: {response.status_code}, Error: {response.errors}")
    return space_values

def report_volumes(fleet_member_name):
    # Go through all volumes on this host and create an array of tags with volume names according to the tagging plan
    volume_set = []
    response = client.get_volumes(context_names=fleet_member_name)
    if response.status_code == 200:
        if debug >= 2:
            print(f"Finding volumes for array {fleet_member_name}")

        for volume in response.items:
            if volume.subtype == 'regular':
                volume_set.append(volume.name)
            else:
                if debug >= 4:
                    print(f'Non-regular volume {volume.name} found - not reporting it.')

        tags = read_volume_tags(fleet_member_name, volume_set)
        space_values = get_volume_space(fleet_member_name, volume_set)

        volumes_by_tag = {}
        volumes_without_tag = []

        # Define headers dynamically based on the first space dictionary encountered
        if space_values:
            first_space = next(iter(space_values.values()))
            headers = ['Date/Time', 'Array', 'Volume'] + list(first_space.keys())
            HEADER_ROWS[0] = headers

        for volume in volume_set:
            tag = tags.get(volume, 'No tag')
            space = space_values.get(volume, {})
            volume_info = {
                'Date/Time': NOW,
                'Array': fleet_member_name,
                'Volume': volume
            }
            # Add the space attributes dynamically
            for key in HEADER_ROWS[0][3:]:  # Skip the first three columns
                volume_info[key] = space.get(key, '')

            if tag == 'No tag':
                volumes_without_tag.append(volume_info)
            else:
                if tag not in volumes_by_tag:
                    volumes_by_tag[tag] = []
                volumes_by_tag[tag].append(volume_info)

        return volumes_by_tag, volumes_without_tag
    else:
        print(f"Failed to retrieve volumes. Status code: {response.status_code}, Error: {response.errors}")
        return {}, []

USER_NAME = ""
TAG_API_TOKEN = ""
FUSION_SERVER = ""
NAMESPACE = ""
SPACE = {}
NOW = ""

# Header rows for the reporting spreadsheet are defined here
HEADER_ROWS = [
    ['Date/Time', 'Array', 'Volume']
]

# Disable certificate warnings
urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)

# Main script
if __name__ == "__main__":
    # Read the Excel file
    tagging_spreadsheet = pd.ExcelFile('STAAS_Tagging.xlsx')
    # Extract global variables from the Fleet sheet
    fleet_df = tagging_spreadsheet.parse('Fleet')

    global_variables = fleet_df.iloc[0].to_dict()

    # Assign global variables
    USER_NAME = os.getenv('PURE_USER_NAME')
    API_TOKEN = os.getenv('PURE_API_TOKEN')
    FUSION_SERVER = global_variables.get('FUSION_SERVER', '')
    NAMESPACE = global_variables.get('NAMESPACE', '')
    TAG_KEY = "chargeback"

    NOW = datetime.now().strftime("%Y-%m-%d %H:%M")

    client = initialise_client(FUSION_SERVER, USER_NAME, API_TOKEN)
    if not client:
        exit(1)
    role = check_purity_role(client, USER_NAME) 
    # Check to see minimum version of 2.38 & array admin privileges for this user
    if not (role == "array_admin" or role == "read_only)"):
        exit(1)
    if not check_api_version(client, 2.38, debug):
        exit(2)

    # Get the arrays for reporting contexts for the nominated fleet
    all_volumes_by_tag = {}
    all_volumes_without_tag = []

    for fleet_member in list_fleets(client):
        volumes_by_tag, volumes_without_tag = report_volumes(fleet_member.member.name)
        for tag, volumes in volumes_by_tag.items():
            if tag not in all_volumes_by_tag:
                all_volumes_by_tag[tag] = []
            all_volumes_by_tag[tag].extend(volumes)
        all_volumes_without_tag.extend(volumes_without_tag)

    # Create or append to the reporting spreadsheet
    file_path = 'STAAS_Reporting.xlsx'
    try:
        if os.path.exists(file_path):
            try:
                book = load_workbook(file_path)
                with pd.ExcelWriter(file_path, engine='openpyxl', mode='a', if_sheet_exists='overlay') as writer:
                    writer.book = book
                    for tag, volumes in all_volumes_by_tag.items():
                        df = pd.DataFrame(volumes)
                        sheet_name = f"Chargeback {tag}"
                        if sheet_name in writer.book.sheetnames:
                            startrow = writer.book[sheet_name].max_row
                            df.to_excel(writer, sheet_name=sheet_name, index=False, header=False, startrow=startrow)
                        else:
                            df.to_excel(writer, sheet_name=sheet_name, index=False)
                    if all_volumes_without_tag:
                        df = pd.DataFrame(all_volumes_without_tag)
                        if 'No Tag' in writer.book.sheetnames:
                            startrow = writer.book['No Tag'].max_row
                            df.to_excel(writer, sheet_name='No Tag', index=False, header=False, startrow=startrow)
                        else:
                            df.to_excel(writer, sheet_name='No Tag', index=False)
            except KeyError as e:
                print(f"KeyError: {e}. The file might be corrupted. Creating a new file.")
                with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
                    for tag, volumes in all_volumes_by_tag.items():
                        df = pd.DataFrame(volumes)
                        df.to_excel(writer, sheet_name=f"Chargeback {tag}", index=False)
                    if all_volumes_without_tag:
                        df = pd.DataFrame(all_volumes_without_tag)
                        df.to_excel(writer, sheet_name='No Tag', index=False)
        else:
            with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
                for tag, volumes in all_volumes_by_tag.items():
                    df = pd.DataFrame(volumes)
                    df.to_excel(writer, sheet_name=f"Chargeback {tag}", index=False)
                if all_volumes_without_tag:
                    df = pd.DataFrame(all_volumes_without_tag)
                    df.to_excel(writer, sheet_name='No Tag', index=False)
    except PermissionError as e:
        print(f"PermissionError: {e}. Please ensure the file is not open in another application.")
```python
# get a list of fleets from this array
def list_fleets(client):
    # Retrieve the list of fleets, then find all of the FlashArrays and volumes associated with the fleet
    response = client.get_fleets_members()
    if response.status_code == 200:
        fleets_members=response.items
    else:
        print(f"Failed to retrieve fleets/members. Status code: {response.status_code}, Error: {response.errors}")
    return(fleets_members)

# Check API version
def check_api_version(client, level, debug=1):
    version = client.get_rest_version()
    
    if float(version) >= level:
        return True
    else:
        if debug >= 1:
            print(f'API Version: {version}')
            print(f"API version needs to support Fusion v{level} at a minimum.")
        return False

# check if the user has the correct permissions (array_admin for tagging, or lesser read/readonly role for reporting)
def check_purity_role(client, USER_NAME, debug=1):
    try:
        response = client.get_admins()
        if response.status_code == 200:
            return "array_admin"
        elif response.status_code == 400:
            print(f'Failed to get admins: {response.errors}')
            return "readonly"
    except PureError as e:
        print(f'Error checking admin level: {e}')
    return False

# make the initial connection to the Fusion server
def initialise_client(FUSION_SERVER, USER_NAME, API_TOKEN):
    if not FUSION_SERVER:
        print("Fusion server not set.")
        return None
    if not USER_NAME:
        print("User name not set.")
        return None
    if not API_TOKEN: 
        print("API token not set.")
        return None
    try:
        client = Client(FUSION_SERVER, username=USER_NAME, api_token=API_TOKEN)
        return client
    except PureError as e:
        print(f"Error initializing client: {e}")
        return None